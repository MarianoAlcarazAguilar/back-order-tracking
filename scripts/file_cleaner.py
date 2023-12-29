import re
import os
import warnings
import calendar
import pandas as pd
import numpy as np
import openpyxl  as xl
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

if __name__ != '__main__':
    from scripts.excel_functions import ExcelFunctions
else:
    from excel_functions import ExcelFunctions



class FileCleaner:
    def __init__(self, main=False, data_dir:str='.'):
        self.ef = ExcelFunctions()
        self.data_dir = data_dir
        if main:
            self.unidades = pd.read_csv(f'../static_data/conversion_unidades.csv', usecols=['sku', 'pza_cja'])
        else:
            self.unidades = pd.read_csv(f'static_data/conversion_unidades.csv', usecols=['sku', 'pza_cja'])

    def __obtener_mes_actual_y_siguiente(self, forzar_fecha: str = None):
        if forzar_fecha is None:
            # Obtener la fecha actual del sistema
            fecha_actual = datetime.now()
        else:
            fecha_actual = pd.to_datetime(forzar_fecha, dayfirst=True)
        
        # Obtener el número del mes actual (1 para enero, 2 para febrero, etc.)
        mes_actual = fecha_actual.month
        
        # Obtener el número del siguiente mes
        mes_siguiente = mes_actual + 1 if mes_actual < 12 else 1

        # Obtener los nombres de los meses en español
        meses_en_espanol = calendar.month_name[mes_actual].lower(), calendar.month_name[mes_siguiente].lower()

        return list(meses_en_espanol)
    
    def __extraer_fecha_from_nombre(self, nombre):
        # Definimos una lista de posibles patrones de fecha que podrían aparecer en el nombre
        patrones_fecha = [
            r'\b\d{1,2}-\d{1,2}-\d{4}\b',   # Formato dd-mm-yyyy
            r'\b\d{1,2}/\d{1,2}/\d{4}\b',   # Formato dd/mm/yyyy
            r'\b\d{4}-\d{1,2}-\d{1,2}\b',   # Formato yyyy-mm-dd
            r'\b\d{4}/\d{1,2}/\d{1,2}\b',   # Formato yyyy/mm/dd
        ]

        # Buscamos los patrones en el nombre y extraemos la primera coincidencia de fecha
        fecha = None
        for patron in patrones_fecha:
            coincidencias = re.findall(patron, nombre)
            if coincidencias:
                fecha = coincidencias[0]
                break

        return fecha
    
    def __set_almacenes_to_cero(self, x, unwanted_almacenes = [136, 17]):
        '''
        Esta función se enciarga de convertir a 0 todos los almacenes especificados
        '''
        almacen = x.almacen
        if almacen in unwanted_almacenes:
            return 0
        else:
            return x.cantidad
        
    def __convierte_a_cajas(self, x, col_names={'unidad_medida':'um', 'cantidad':'cantidad'}):
        '''
        Esta función convierte todasl as piezas en cajas
        '''
        um = x[col_names['unidad_medida']]
        cantidad = x[col_names['cantidad']]
        if um == 'PZ':
            resp = cantidad / x.pza_cja
        else:
            resp = cantidad
        resp = np.ceil(resp)
        return resp
    
    def __clean_planta_name(self, x, index_pos=-1):
        '''
        Esta función limpia los nombres de las plantas para que no sean tan largos
        '''
        name = x.planta
        try:
            return name.split()[index_pos].lower()
        except:
            print(f'Algo no está bien con el nombre: {name}')
            return None
    
    def __obtener_posicion_por_indice(self, diccionario, indice):
        diccionario_salida = {}

        for llave, lista in diccionario.items():
            # Verificamos si el índice es válido para la lista
            if indice < len(lista):
                diccionario_salida[llave] = lista[indice]
            else:
                # Si el índice está fuera del rango, seleccionamos el último elemento disponible
                diccionario_salida[llave] = lista[-1]

        return diccionario_salida
    
    def __convertir_fechas_a_meses_en_ingles(slef, diccionario):
        diccionario_meses_en_ingles = {}

        for fecha, valores in diccionario.items():
            # Verificamos si la llave es del tipo datetime
            if isinstance(fecha, datetime):
                # Obtenemos el nombre del mes en inglés
                nombre_mes = calendar.month_name[fecha.month].lower()
                new_key = f'{nombre_mes}_{fecha.year}'
                # Agregamos la entrada al nuevo diccionario
                diccionario_meses_en_ingles[new_key] = valores
            else:
                diccionario_meses_en_ingles[fecha] = valores

        return diccionario_meses_en_ingles
    
    def __hex_to_RGB(self, hex_str):
        """ #FFFFFF -> [255,255,255]"""
        #Pass 16 to the integer function for change of base
        return [int(hex_str[i:i+2], 16) for i in range(1,6,2)]

    def __get_color_gradient(self, c1, c2, n):
        """
        Given two hex colors, returns a color gradient
        with n colors.
        """
        assert n > 1
        c1_rgb = np.array(self.__hex_to_RGB(c1))/255
        c2_rgb = np.array(self.__hex_to_RGB(c2))/255
        mix_pcts = [x/(n-1) for x in range(n)]
        rgb_colors = [((1-mix)*c1_rgb + (mix*c2_rgb)) for mix in mix_pcts]
        return ["".join([format(int(round(val*255)), "02x") for val in item]) for item in rgb_colors]
    
    def __encontrar_numeros_en_cadena(self, cadena):
        numeros_encontrados = re.findall(r'-?\d+(?:\.\d+)?', cadena)
        return numeros_encontrados
    
    def __encontrar_maximo_datetime(self, lista_datetime) -> int:
        max_datetime, max_index = None, None

        for i, dt in enumerate(lista_datetime):
            if max_datetime is None or dt > max_datetime:
                max_datetime = dt
                max_index = i

        return max_index
    
    def __encontrar_archivo_mas_reciente(self, nombres_archivos: list, dia_primero: bool=True):
        '''
        Esta función recibe una lista de nombres de fechas que contienen una fecha,
        y regresa el nombre de archivo con la fecha más reciente. 
        Funciona bien con los nombres de inventarios y transferencias.
        Para forecasts names usar encontrar_forecast_mas_reciente()
        '''
        fechas = []
        
        for file in nombres_archivos:
            numeros_encontrados = self.__encontrar_numeros_en_cadena(file)
            
            if len(numeros_encontrados) != 3:
                print(f'No se pudo identificar la fecha en: {file}')
                continue
                
            fecha_encontrada = '/'.join(numeros_encontrados)
            fechas.append(pd.to_datetime(fecha_encontrada, dayfirst=dia_primero))
            
        max_date_index = self.__encontrar_maximo_datetime(fechas)
        
        if max_date_index is not None: # Si es válido el índice lo regresamos
            return nombres_archivos[max_date_index]
        
        print('No se encontraron fechas válidas en ningún archivo')
        return None
    
    def __encontrar_forecast_mas_reciente(self, lista_archivos: list):
        '''
        Esta función recibe nombres de archivos de forecasts y regresa 
        aquel que contenga el mes y año más recientes.
        '''
        max_fecha = None
        archivo_mas_reciente = None

        for nombre_archivo in lista_archivos:
            match = re.findall(r'(\w+)_(\w+)_(\d{4})_(\w+)_(\d{4})', nombre_archivo)
            if match:
                mes_inicio = match[0][1]
                ano_inicio = int(match[0][2])
                mes_fin = match[0][3]
                ano_fin = int(match[0][4])

                fecha_inicio = datetime.strptime(f"{mes_inicio} {ano_inicio}", "%B %Y")
                fecha_fin = datetime.strptime(f"{mes_fin} {ano_fin}", "%B %Y")

                if max_fecha is None or fecha_fin > max_fecha:
                    max_fecha = fecha_fin
                    archivo_mas_reciente = nombre_archivo

        return archivo_mas_reciente
    
    def __asigna_color_celda(self, x, color_inicio='#caf0f8', color_fin='#023e8a', n_colores=70):
        # Se tiene una celda que se llama porcentaje
        #color_inicio='#ebf4f5'
        color_fin='#014f86'
        if x.porcentaje >= 0:
            return 'ffffff'
        
        colores = self.__get_color_gradient(color_inicio, color_fin, n_colores)
        color_index = int(np.round(x.porcentaje*-n_colores)-1)
        if color_index < 0:
            color_index = 0
        return colores[color_index]
    
    def __asigna_color_urgencia(self, x, color_inicio='#fcbf49', color_fin='#d62828', n_colores=3):
        if x.porcentaje >= 0:
            return 'ffffff'
        
        colores = self.__get_color_gradient(color_inicio, color_fin, n_colores)
        color_index = int(np.round(x.porcentaje*-n_colores)-1)
        if color_index < 0:
            color_index = 0
        return colores[color_index]
    
    def __obtener_contraste(self, x):
        color_hex = x.color
        r = int(color_hex[0:2], 16)
        g = int(color_hex[2:4], 16)
        b = int(color_hex[4:6], 16)
        
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        
        if luminance > 0.5:
            return "000000"
        else:
            return "ffffff"
    
    def __ajustar_anchura_columnas(self, ws):
        for columna in ws.columns:
            max_length = 0
            columna_letra = get_column_letter(columna[0].column)
            for celda in columna:
                try:
                    if len(str(celda.value)) > max_length:
                        max_length = len(celda.value)
                except:
                    pass
            ajuste_anchura = (max_length + 2)
            ws.column_dimensions[columna_letra].width = ajuste_anchura

    def __get_latest_files(self): # def __get_latest_files(self, current_dir='.')
        current_dir = self.data_dir
        dir_forecast = f'{current_dir}/transformed_data/forecasts/'
        dir_inventarios = f'{current_dir}/transformed_data/inventarios/'
        dir_transfers = f'{current_dir}/transformed_data/transferencias/'

        latest_forecast = self.__encontrar_forecast_mas_reciente(os.listdir(dir_forecast))
        latest_inventario = self.__encontrar_archivo_mas_reciente(os.listdir(dir_inventarios))
        latest_transfer = self.__encontrar_archivo_mas_reciente(os.listdir(dir_transfers))

        forecast = pd.read_excel(f'{dir_forecast}/{latest_forecast}').drop(['descripcion'], axis=1)
        inventario = pd.read_excel(f'{dir_inventarios}/{latest_inventario}')[['sku', 'inventario']]
        transfers = pd.read_excel(f'{dir_transfers}/{latest_transfer}')[['sku', 'total']]

        data_frames = (forecast, inventario, transfers)
        file_names = (latest_forecast, latest_inventario, latest_transfer)

        return file_names, data_frames
    
    def __obtener_fecha_from_filename(self, filename, day_first=True):
        """
        Esta función recibe un filename con fecha y regresa un datetime de él
        """
        numeros = self.__encontrar_numeros_en_cadena(filename)
        fecha_encontrada = '/'.join(numeros)
        return pd.to_datetime(fecha_encontrada, dayfirst=day_first)
    
    def __get_date_from_transfer_filename(self, filename, day_first=True):
        """
        Los nombres de los reportes de transferencias siempre tienen el siguente foramto:
        Reporte T2 dd-mm-yyyy.xlsx
        """
        filename = self.eliminar_parentesis_duplicados(filename)
        fecha_encontrada = filename.split('.')[0].split(' ')[-1]
        return pd.to_datetime(fecha_encontrada, dayfirst=day_first)
    
    def clean_inventory_file(self, file, dir_location: str, saving_name=None, return_data=True):
        '''
        Esta función recibe un archivo, ya sea ubicación para abrirlo o bytes y lo limpia.
        dir_location: el directorio donde lo va a guardar
        '''
        # Definimos parámetros necesarios
        sheet_name = 'Rep Prod Terminado Ubi Chh Mx'
        wanted_cols = ['CLAVE', 'Descripción', 'Total x Producto']
        column_output_names = ['sku', 'descripcion', 'inventario']

        # Creamos el objeto de funciones de excel
        ef = self.ef

        # Abrimos el archivo pedido
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            wb = xl.load_workbook(file, data_only=True)
        ws = wb[sheet_name]
        fecha = self.__extraer_fecha_from_nombre(file.name)

        # Encontramos la fila con el nombre de las columnas
        row_with_col_names = ef.find_value(ws, 1, text='CLAVE')[0]
        last_row_with_data = ef.encontrar_ultima_fila_con_datos(ws, 1, 5)

        # Encontramos los nombres de las columnas en esa fila
        dict_cols = ef.get_columns_on_row(ws, row_with_col_names, num_nones_consecutivos=3)

        # Sacamos los datos de las columnas deseadas
        extracted_data = ef.get_data_on_columns(
            ws, 
            row_with_col_names+1, 
            dict_cols, 
            ending_row=last_row_with_data, 
            cols_to_extract=wanted_cols,
            output_columns=column_output_names
        )

        # Procesamos un poco los datos
        extracted_data = (extracted_data
        .assign(
            fecha=fecha,
            inventario=lambda x: pd.to_numeric(x.inventario, errors='coerce')
        )
        [['fecha', 'sku', 'descripcion', 'inventario']]
        )
        
        # Guardamos el archivo
        if saving_name is None:
            saving_name = f'inventario_{fecha.replace("-", "_")}'
            
        bytes_data, filename = ef.save_and_download_excel_file(
            extracted_data,
            dir_location=dir_location,
            file_name=saving_name,
            sheet_name='Inventario',
            n_cols_to_bold=3,
            return_data=return_data
        )
        
        if return_data:
            return bytes_data, filename
        
        print(f'{saving_name} was updated succesfully')

    def clean_transferencias(self, file, dir_location: str, return_data=True):
        '''
        Esta fucnión se encuarga de toda la funcionalidad
        '''
        ef = ExcelFunctions()
        
        df_transferencias = pd.read_excel(file)
        
        columnas_necesarias = ['Nombre venta', 'Articulo', 'Descripción', 'Cantidad', 'UM', 'Almacén']
        fecha = self.__extraer_fecha_from_nombre(file.name)    
           
        processed_transfers = (df_transferencias
        [columnas_necesarias]
        .rename(columns={'Nombre venta':'planta', 'Articulo':'sku', 'Cantidad':'cantidad', 'UM': 'um', 'Almacén':'almacen', 'Descripción':'descripcion'})
        .merge(self.unidades, on='sku', how='left')
        .assign(
            c_=lambda df: df.apply(lambda x: self.__set_almacenes_to_cero(x), axis=1), # Convertimos a 0 los almacenes no deseados,
            cantidad=lambda df: df.apply(lambda x: self.__convierte_a_cajas(x, col_names={'unidad_medida':'um', 'cantidad':'c_'}), axis=1), # Convertimos todo a cajas
            fecha=fecha, # La voy a dejar en formato string
            planta=lambda df: df.apply(self.__clean_planta_name, axis=1)
        ) 
        .drop(['c_', 'pza_cja', 'um'], axis=1)
        .pivot_table(index=['fecha', 'sku', 'descripcion'], columns='planta', values='cantidad', aggfunc='sum') # Hacemos el reporte por planta
        .assign(total=lambda x: x.sum(axis=1)) # Sumamos el total
        .reset_index()
        .assign(sku=lambda x: x.sku.astype('str'))
        .sort_values('sku')
        )
        
        # Guardamos el archivo generado en formato excel bonito
        saving_name = f"transferencias_plantas_{fecha.replace('-','_')}"
        byte_data, filename = ef.save_and_download_excel_file(
            df=processed_transfers,
            dir_location=dir_location,
            file_name=saving_name,
            sheet_name='Transferncias',
            n_cols_to_bold=3,
            return_data=return_data
        )
        if return_data:
            return byte_data, filename
        
        print(f'{saving_name} updated succesfully')

    def clean_forecast_file(self, file, dir_location: str, file_name=None, return_data=True):
        '''
        Esta función recibe un archivo, ya sea ubicación para abrirlo o bytes y lo limpia.
        dir_location: el directorio donde lo va a guardar
        '''
        # Aquí están las variables necesarias
        ef = ExcelFunctions()

        # Nombres necesarios
        sheet_name = 'Resumen Pzs, Cjs, Kg-Lt'
        first_col_name = 'Item'
        wanted_cols = ['Item', 'Description']
        output_col_names = ['sku', 'descripcion']

        # Abrimos el archivo
        wb = xl.load_workbook(file, data_only=True)
        ws = wb[sheet_name]
        
        # Encontramos las dimensiones necesarias
        row_with_col_names = ef.find_value(ws, 1, first_col_name)[0]
        last_row = ef.encontrar_ultima_fila_con_datos(ws, 1, row_with_col_names)

        # Traemos los nombres de las columnas
        # Primero tenemos que ver a partir de qué columna están las cajas
        col_cajas = ef.get_columns_on_row(ws, row_with_col_names-1, num_nones_consecutivos=20)['CAJAS']
        dict_cols = ef.get_columns_on_row(ws, row_with_col_names)
        
        # Sacamos los nombres de las columnas repetidas
        # Convertimos a nombres de meses y años
        # Tomamos la segunda entrada si disponible
        otro_dict_cols = ef.get_columns_on_row_repeated_names(ws, row_with_col_names) # NEW NEW NEW
        otro_dict_cols = self.__convertir_fechas_a_meses_en_ingles(otro_dict_cols)# new new new
        otro_dict_cols = self.__obtener_posicion_por_indice(otro_dict_cols, 1) # new new new

        # Tenemos que agregar los nombres de los siguientes dos meses para la extraccion
        meses = self.__obtener_mes_actual_y_siguiente()
        if "december" in meses and "january" in meses:
            this_year = datetime.now().year
            otros_meses = [f"december_{this_year}", f"january_{this_year+1}"]
        else:
            # Le agregamos el año actual a los meses a extraer
            # Aquí hay un problema en diciembre: enero no te lo manda al siguiente año
            otros_meses = [f'{mes}_{datetime.now().year}' for mes in meses] # NEW NEW NEW
        otro_cols_to_extract = wanted_cols + otros_meses # NEW NEW NEW

        for i, mes in enumerate(meses):
            dict_cols[mes] = col_cajas+i

        col_names = output_col_names + meses
        extracted_data = (ef.get_data_on_columns(
            ws, 
            row_with_col_names+1, 
            otro_dict_cols, # new new new 
            last_row, 
            otro_cols_to_extract, # new new new 
            col_names
        )
        )

        # Limpiamos los datos extraídos
        for mes in meses:
            extracted_data[mes] = np.round(pd.to_numeric(extracted_data[mes]))
            
        # Agregamos el año actual a los nombres de los meses
        extracted_data = (extracted_data
        .set_index(['sku', 'descripcion'])
        .add_suffix(f' {datetime.now().year}')
        .reset_index()
        )
            
        if file_name is None:
            file_name = f'forecast_{"_".join(otros_meses)}'
            
        # Guardamos el archivo
        data_bytes, filename = ef.save_and_download_excel_file(
            extracted_data, 
            dir_location, 
            file_name,
            sheet_name = 'Forecast', 
            n_cols_to_bold=2,
            return_data=return_data
        )
        if return_data:
            return data_bytes, filename
        
        print(f'{file_name} was updated succesfully')

    def genera_reporte_back_order_especifico(self, files:dir):
        '''
        files es un diccionario de la forma {'forecast':filename, 'inventario':filename, 'transfers':filename}
        '''
        current_dir = self.data_dir
        ef = ExcelFunctions()

        dir_back_orders = f'{current_dir}/transformed_data/back_orders'
        dir_forecast = f'{current_dir}/transformed_data/forecasts/'
        dir_inventarios = f'{current_dir}/transformed_data/inventarios/'
        dir_transfers = f'{current_dir}/transformed_data/transferencias/'

        forecast = pd.read_excel(f'{dir_forecast}/{files["forecast"]}').drop(['descripcion'], axis=1)
        inventario = pd.read_excel(f'{dir_inventarios}/{files["inventario"]}')[['sku', 'inventario']]
        transfers = pd.read_excel(f'{dir_transfers}/{files["transfers"]}')[['sku', 'total']]

        clasificaciones = pd.read_excel(f'static_data/clasificacion_abc.xlsx').drop(['planta'], axis=1)

        # Transformamos el forecast y guardamos los nombres de las columnas
        forecast = (forecast
         .set_index('sku')
         .assign(total_forecast=lambda df: df.sum(axis=1))
        )

        # Hacmeos el nombre, se toma como base el nombre del archivo de inventario
        date_numbers_inventario = "_".join(self.__encontrar_numeros_en_cadena(files['inventario']))
        file_name = f"reporte_back_order_{date_numbers_inventario}"

        # Sumamos las transferencias y los inventarios por producto
        output_order = ['sku', 'descripcion', 'clasificacion', 'u', 'inventario', 'transferencias'] + list(forecast.columns) + ['existencias']

        reporte = (transfers
         .rename(columns={'total':'transferencias'})
         .merge(inventario, on='sku', how='outer')
         .merge(forecast, on='sku', how='left')
         .merge(clasificaciones, on='sku', how='left')
         # Quitamos los productos que no queremos
         .query('inventario > 0 and (clasificacion == "A" or clasificacion == "B") and total_forecast > 0') #
         # Sumamos inventarios y transferencias
         .fillna({'transferencias':0})
         .assign(
             existencias=lambda x: x.transferencias + x.inventario,
             transferencias=lambda x: x.transferencias.replace(0, np.nan),
             porcentaje=lambda x: x.existencias / x.total_forecast - 1,
             u=None
         )
         .sort_values('porcentaje')
         [output_order]
        )

        # El color de este criterio es escacez relativa
        colores = (reporte
         .assign(
             porcentaje=lambda x: x.existencias / x.total_forecast - 1, # Mientras más cercano a -1 es más urgente 
             color=lambda x: x.apply(self.__asigna_color_celda, axis=1),
             texto=lambda x: x.apply(self.__obtener_contraste, axis=1),
             urgencia=lambda x: x.apply(self.__asigna_color_urgencia, axis=1)
         )
         [['color', 'texto', 'urgencia']]
        )
        
        n_cols_to_bold = 3

        _, filename = ef.save_and_download_excel_file(
            reporte, 
            dir_back_orders, 
            file_name,
            sheet_name='Reporte Back Order',
            n_cols_to_bold=n_cols_to_bold,
            return_data=True
        )
        
        # Ahora tenemos que colorear del color específico
        # Abrimos el archivo creado
        wb = xl.load_workbook(f'{dir_back_orders}/{filename}', data_only=True)
        ws = wb.active

        rows, cols = reporte.shape

        for row, colors in enumerate(colores.values):
            row += 2
            color_celda, color_texto, color_urgencia = colors
            fill_color = PatternFill(start_color=f'{color_celda}', end_color=f'{color_celda}', fill_type='solid')
            fill_urgencia = PatternFill(start_color=f'{color_urgencia}', end_color=f'{color_urgencia}', fill_type='solid')
            text_color = Font(color=color_texto)
            for col in range(n_cols_to_bold+2, cols+1):
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).font = text_color
            ws.cell(row=row, column=n_cols_to_bold+1).fill = fill_urgencia

        # Ajustamos la anchura del resto de las columnas
        self.__ajustar_anchura_columnas(ws)

        wb.save(f'{dir_back_orders}/{filename}')
        return {
            'Forecast': files["forecast"],
            'Inventario': files["inventario"],
            'Transferencias': files["transfers"],
            'filename': filename
        }

    def genera_reporte_back_order(self): # def genera_reporte_back_order(self, current_dir='.'):
        current_dir = self.data_dir
        ef = ExcelFunctions()

        dir_back_orders = f'{current_dir}/transformed_data/back_orders'

        file_names, data_frames = self.__get_latest_files() # Extraemos y leemos los datos más recientes
        forecast_filename, inventario_filename, transfer_filename = file_names # Los sacamos de la tupla
        forecast, inventario, transfers = data_frames

        clasificaciones = pd.read_excel(f'static_data/clasificacion_abc.xlsx').drop(['planta'], axis=1)

        # Transformamos el forecast y guardamos los nombres de las columnas
        forecast = (forecast
         .set_index('sku')
         .assign(total_forecast=lambda df: df.sum(axis=1))
        )

        file_name = f"reporte_back_order_{datetime.now().date().strftime('%d_%m_%Y')}"
        
        # Sumamos las transferencias y los inventarios por producto
        output_order = ['sku', 'descripcion', 'clasificacion', 'u', 'inventario', 'transferencias'] + list(forecast.columns) + ['existencias']

        reporte = (transfers
         .rename(columns={'total':'transferencias'})
         .merge(inventario, on='sku', how='outer')
         .merge(forecast, on='sku', how='left')
         .merge(clasificaciones, on='sku', how='left')
         # Quitamos los productos que no queremos
         .query('inventario > 0 and (clasificacion == "A" or clasificacion == "B") and total_forecast > 0') #
         # Sumamos inventarios y transferencias
         .fillna({'transferencias':0})
         .assign(
             existencias=lambda x: x.transferencias + x.inventario,
             transferencias=lambda x: x.transferencias.replace(0, np.nan),
             porcentaje=lambda x: x.existencias / x.total_forecast - 1,
             u=None
         )
         .sort_values('porcentaje')
         [output_order]
        )

        # El color de este criterio es escacez relativa
        colores = (reporte
         .assign(
             porcentaje=lambda x: x.existencias / x.total_forecast - 1, # Mientras más cercano a -1 es más urgente 
             color=lambda x: x.apply(self.__asigna_color_celda, axis=1),
             texto=lambda x: x.apply(self.__obtener_contraste, axis=1),
             urgencia=lambda x: x.apply(self.__asigna_color_urgencia, axis=1)
         )
         [['color', 'texto', 'urgencia']]
        )
        
        n_cols_to_bold = 3

        _, filename = ef.save_and_download_excel_file(
            reporte, 
            dir_back_orders, 
            file_name,
            sheet_name='Reporte Back Order',
            n_cols_to_bold=n_cols_to_bold,
            return_data=True
        )
        
        # Ahora tenemos que colorear del color específico
        # Abrimos el archivo creado
        wb = xl.load_workbook(f'{dir_back_orders}/{filename}', data_only=True)
        ws = wb.active

        rows, cols = reporte.shape

        for row, colors in enumerate(colores.values):
            row += 2
            color_celda, color_texto, color_urgencia = colors
            fill_color = PatternFill(start_color=f'{color_celda}', end_color=f'{color_celda}', fill_type='solid')
            fill_urgencia = PatternFill(start_color=f'{color_urgencia}', end_color=f'{color_urgencia}', fill_type='solid')
            text_color = Font(color=color_texto)
            for col in range(n_cols_to_bold+2, cols+1):
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).font = text_color
            ws.cell(row=row, column=n_cols_to_bold+1).fill = fill_urgencia

        # Ajustamos la anchura del resto de las columnas
        self.__ajustar_anchura_columnas(ws)

        wb.save(f'{dir_back_orders}/{filename}')
        return {
            'Forecast': forecast_filename,
            'Inventario': inventario_filename,
            'Transferencias': transfer_filename,
            'filename': filename
        }
    
    def ordena_lista_nombres_con_fecha(self, filenames:list, reverse=True):
        return sorted(filenames, key=self.__obtener_fecha_from_filename, reverse=reverse)
    
    def ordena_lista_transferencias(self, filenames: list, reverse=True):
        return sorted(filenames, key=self.__get_date_from_transfer_filename, reverse=True)
    
    def eliminar_parentesis_duplicados(self, nombre_archivo):
        # Utilizamos una expresión regular para buscar y eliminar los paréntesis y su contenido.
        nuevo_nombre = re.sub(r'\s*\(\d+\)', '', nombre_archivo)
        return nuevo_nombre

    def encontrar_numeros_en_cadena(self, cadena):
        numeros_encontrados = re.findall(r'-?\d+(?:\.\d+)?', cadena)
        return numeros_encontrados

if __name__ == '__main__':
    fc = FileCleaner(main=True)

    fc.clean_transferencias(
        '../raw_data/Reporte T2 31-07-2023.xlsx', 
        '../transformed_data',
        'Reporte T2 31-07-2023.xlsx',
        return_data=False
    )

    fc.clean_inventory_file(
        '../raw_data/DISTRIBUCIÓN Y ABASTO 31-07-2023.xlsx',
        '../transformed_data',
        return_data=False
    )

    fc.clean_forecast_file(
        '../raw_data/Forecast Estadístico Colaborado Julio 2023 Nacional (1).xlsx',
        '../transformed_data',
        return_data=False
    )